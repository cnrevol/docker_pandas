"""
文件做成，输出类
"""
from datetime import datetime
import re
import os
import traceback
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import json

from utils.db_util import Database
from utils.file_util import FileUtil
from utils.comm_util import CommUtil
from exceptions import CustomException

class OutputFileAction:
    """
    文件输出类定义
    """

    POST_FILE = "post"
    ALOC_FILE = "aloc"
    POST_T_FILE = "post2"

    C_SRC = "src_col"
    C_DST ="dst_col"
    C_IDX ="col_idx"

    STS_PD="data_posted"
    # bank code define

    # bank_code_map={
    #     "SGHSBCSGD":"H02",
    #     "SGHSBCUSD":"H08",
    #     "SGHSBCTHB":"H11",
    #     "SGHSBCPHP":"H07",
    #     "SGHSBCVND":"H05",
    #     "SGBOAKRW":"H03",
    #     "MYHSBCMYR":"M01",
    #     "MYMBBMYR":"M10",
    #     "MYM2UMYR":"M11",
    #     "HKHSBCHKD":"H01",
    #     "HKHSBCUSD":"H14",
    #     "HKBOAHKD":"B01",
    #     "HKBOAUSD":"B02",
    #     "AUHSBCAUD":"HS1",
    #     "NZHSBCNZD":"HS2",
    #     "NZBOANZD":"CB1",
    #     "TWHSBCUSD":"H08",
    #     "TWHSBCTWD":" ",
    #     "INBOAINR":"BOA"
    # }

    doc_types = ('SIN', 'SLI', 'SRP')
    # user_confirmed 

    def  __init__(self, name: str,logger = None) -> None:
        self.name = name
        self.logger = logger
        self.db = Database(logger = logger)
        self.file_util = FileUtil(logger=logger)
        self.comm_util = CommUtil(logger=logger)
    

    def output_alloc_data(self, region, action_user="system"):
        """
        allocate 销账数据输出到文件
        """
        self.logger.info(f"Allocation file output start Region: {region} ")
        try:
            bz_date = self.comm_util.get_daytime_string()
            aloc_action_name = f"{region}_Allocation_file_output"
            self.comm_util.update_alocfile_outputing_status(
                region=region,
                action_name=aloc_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["16"],
                bl_message=f"Allocation file output action on Region:{region} start.",
                ac_user=action_user,
            )
            table_name = "bl_aging_allocate"

            dest_df = self.get_aloc_data(region)

            if not dest_df.empty:
                # 取得配置表定义
                def_df = self.get_out_file_def(region, self.ALOC_FILE)
                out_fields = def_df.loc[0, "out_field_define"]
                mapping_list = self.raw_column_define(out_fields)

                # 生成 detail 和 header 数据
                detail_df = self.create_alloc_detail_data(dest_df, mapping_list)
                header_df = self.create_alloc_header_data(dest_df, region, action_user)

                # 输出到模板文件
                self.out_to_alloc_template_file(region, header_df, detail_df, self.file_util.ALOC_TEMPLATE)

                dest_df = dest_df.fillna({"update_time": datetime.now()})
                dest_df['aloc_date'] = datetime.now()
                self.update_aloc_data(table_name, dest_df)
                self.process_and_update_cash_data(dest_df, region)
                self.logger.debug(f"Allocation data db status updated.{region} ")
            else:
                self.logger.warning("There is no Allocation data to output.")

            self.comm_util.update_alocfile_outputed_status(
                region=region,
                action_name=aloc_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["17"],
                bl_message=f"Allocation file output action on Region:{region} finished .",
                ac_user=action_user,
            )
        except (CustomException, Exception) as e:
            self.logger.error(
                e.message if isinstance(e, CustomException) else str(e.args)
            )
            self.logger.error(traceback.format_exc())
            self.comm_util.update_alocfile_outputed_status(
                region=region,
                action_name=aloc_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["20"],
                bl_message=f"Allocation file output  action on Region:{region} Failed .",
                ac_user=action_user,
            )
        self.logger.info(f"Allocation file output Successfully Region: {region} ")

    def output_post_data(self, region, action_user="system"):
        """
        post数据输出到文件

        """
        if region == "CN":
            self.output_post_data_cn_csv(region, action_user)
            return
    
        self.logger.info(f"Post file output start Region: {region} ")
        try:
            # 更新Action状态日志
            bz_date = self.comm_util.get_daytime_string()
            post_action_name = f"{region}_post_file_output"
            self.comm_util.update_file_outputing_status(
                region=region,
                action_name=post_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["6"],
                bl_message=f"Post file output action on Region:{region} start.",
                ac_user=action_user,
            )

            table_name = "bl_bank_statement"

            # 取得post对象数据
            dest_df = self.get_post_data(region)
            if not dest_df.empty:
                # 处理IN数据，判断doctype类型SCQ，SBT，A60
                # dest_df = self.set_df_doctype(dest_df,region)
                # 处理asw_text
                dest_df = self.fix_data_miss(dest_df,region)
                # 处理document number
                dest_df, define_table_df = self.get_document_number(dest_df, region)
                # 设定bankcode
                dest_df = self.set_bank_code(dest_df,region)

                # 创建合计数据DataFrame
                summary_df = self.create_post_summary_data(dest_df, region, action_user)
                
                # 创建详细数据DataFrame
                detail_df = self.create_post_detail_data(dest_df, region)

                # 输出到模板文件（包含合计和详细两个sheet）
                self.out_to_post_template_file(region=region, summary_df=summary_df, detail_df=detail_df, file_type=self.file_util.POST_TEMPLATE)

                dest_df = dest_df.fillna(
                    {"credit_amount": 0, "debit_amount": 0, "update_time": datetime.now()}
                )
                # 更新数据标识
                # self.update_post_status(table_name, dest_df)
                # 更新状态，转存到历史表 
                self.update_bank_data(table_name, dest_df)

                # 最新的document number更新到DB
                self.db.upsert_table(define_table_df, "sc_sap_document_define")
                
                # 保存入账结果数据，报表用
                self.update_cash_data(dest_df)
                
                self.logger.info(f"Post file output Successfully Region: {region} ")
            else:
                self.logger.error("There is no data to post")
                # raise CustomException(f"There is no data to post")
        
            # 更新Action状态日志
            self.comm_util.update_file_outputed_status(
                region=region,
                action_name=post_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["7"],
                bl_message=f"Post file output action on Region:{region} finished.",
                ac_user=action_user,
            )
        except (CustomException,Exception) as e:
            self.logger.error(
                e.message if isinstance(e, CustomException) else str(e.args)
            )
            self.logger.error(traceback.format_exc())
            self.comm_util.update_file_outputed_status(
                region=region,
                action_name=post_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["11"],
                bl_message=f"Post file output  action on Region:{region} Failed .",
                ac_user=action_user,
            )
        self.logger.info(f"Post file output  Finished. Region: {region} ")

    def out_to_file(self,region,outfile_df,file_type,filename="",header=False):
        """
        输出到文件
        """
        # 取得输出文件模板
        post_file_writer, out_file_path, item_name = self.file_util.open_template(
            region, file_type,filename
        )
        # 打印数据到模板
        self.print_data(
            region=region, bank="", writer=post_file_writer, df=outfile_df,header=header
        )
        # 云存储
        self.move_file_to_cos(item_name, out_file_path)

    def out_to_postfile(self,region,outfile_df,file_type):
        """
        输出post数据文件
        CN时，按照类型份文件输出
        """
        if region=="CN":
            df_shunfeng, df_alipay, df_ccb = self.split_dataframe(outfile_df)
            self.out_to_file(region,df_shunfeng,file_type,"COD")
            self.out_to_file(region,df_alipay,file_type,"ALIPAY")
            self.out_to_file(region,df_ccb,file_type,"CCB")
        else:
            self.out_to_file(region,outfile_df,file_type,"Posting")

    def create_post_summary_data(self, dest_df, region, action_user):
        """
        创建post合计数据DataFrame
        按照银行、币种分组，计算合计金额和记录数量
        """
        # 取得通用账户ID

        commonid = self.get_common_cid(region)

        # 确保必要的列存在，使用数据表中的实际列名
        required_columns = ['bank_branch_name', 'currency', 'amount']
        for col in required_columns:
            if col not in dest_df.columns:
                self.logger.warning(f"Column {col} not found in dest_df, adding default value")
                if col == 'amount':
                    dest_df[col] = 0
                else:
                    dest_df[col] = ''

        # 按银行和币种分组计算合计
        summary_data = []
        for (bank, currency), group in dest_df.groupby(['bank_branch_name', 'currency']):
            # 计算合计金额（使用amount列）
            total_amount = group['amount'].sum()
            total_count = len(group)
            
            # 获取bank_code（如果存在的话）
            bank_code = ''
            if 'bank_code' in group.columns:
                bank_code = group['bank_code'].iloc[0] if not group['bank_code'].isna().all() else ''
            
            summary_row = {
                'Posting Date': datetime.now().strftime("%d%m%y"),
                'Region': region,
                'Bank': bank,
                'Currency': currency,
                'Bank Code': bank_code,
                'Total Count': total_count,
                'Total Amount': total_amount,
                'Suspense ID': commonid,  # 暂时为空，需要从外部取得
                'User': action_user  # 暂时为空，需要从外部取得
            }
            summary_data.append(summary_row)
        
        summary_df = pd.DataFrame(summary_data)
             
        return summary_df

    def create_post_detail_data(self, dest_df, region):
        """
        创建post详细数据DataFrame
        使用配置定义而不是硬编码的字段映射
        """
        # 从配置表取得需要输出的列名
        def_df = self.get_out_file_def(region, self.POST_FILE)
        out_fields = def_df.loc[0, "out_field_define"]
        mapping_list = self.raw_column_define(out_fields)
        
        # 使用配置定义创建detail_df
        detail_df = pd.DataFrame()
        for mapping in mapping_list:
            src = mapping[self.C_SRC]
            dst = mapping[self.C_DST]
            if src and src in dest_df.columns:
                detail_df[dst] = dest_df[src]
            elif src == "0":
                detail_df[dst] = 0
            else:
                detail_df[dst] = ""
        
        # 特殊处理：Bank Value Date 格式化
        if "Bank Value Date" in detail_df.columns:
            detail_df["Bank Value Date"] = pd.to_datetime(detail_df["Bank Value Date"], errors="coerce").dt.strftime("%d%m%y")
        
        # 特殊处理：System Amount 列，设定为0
        if "System Amount" in detail_df.columns:
            detail_df["System Amount"] = 0
        
        # 特殊处理：Text 列截取前30字符
        if "Text" in detail_df.columns:
            detail_df["Text"] = detail_df["Text"].astype(str).str[:30]

        sort_columns = ["Region", "Bank", "Currency", "Credit Amount"]
        existing_sort_columns = [col for col in sort_columns if col in detail_df.columns]
        if existing_sort_columns:
            # 定义升降序规则，前三个升序，最后一个降序
            ascending_flags = [True, True, True, False][:len(existing_sort_columns)]
            detail_df = detail_df.sort_values(
                by=existing_sort_columns, ascending=ascending_flags
            ).reset_index(drop=True)

        return detail_df

    def process_text_field(self, dest_df):
        """
        处理Text字段的专用方法
        未来可以添加具体的设定逻辑
        """
        # 暂时返回空字符串，未来可以添加具体的处理逻辑
        return ''

    def out_to_post_template_file(self, region, summary_df, detail_df, file_type):
        """
        输出post数据到模板文件，包含合计和详细两个sheet
        """
        # 取得输出文件模板
        post_file_writer, out_file_path, item_name = self.file_util.open_template(
            region, file_type,"Posting"
        )
        
        try:
            # 输出合计数据到Posting Header sheet
            summary_df.to_excel(post_file_writer, sheet_name='Posting Header', startrow=0, header=True, index=False)
            
            # 输出详细数据到Posting Detail sheet
            detail_df.to_excel(post_file_writer, sheet_name='Posting Detail', startrow=0, header=True, index=False)
            
            self.logger.info(f"Post template file created successfully. Summary rows: {len(summary_df)}, Detail rows: {len(detail_df)}")
            
        except Exception as e:
            self.logger.error(f"Error creating post template file: {e}")
            raise CustomException(f"Error creating post template file: {e}")
        finally:
            post_file_writer.save()
        
        # 云存储
        self.move_file_to_cos(item_name, out_file_path)

    def split_dataframe(self,outfile_df):
        """
        根据 'trans_type' 列的值将 DataFrame 分成三个 DataFrame。
        
        Args:
            outfile_df (pd.DataFrame): 输入的 DataFrame，包含 'trans_type' 列。
            
        Returns:
            tuple: 包含三个 DataFrame，分别对应 'CCB' 'shunfeng'， 'alipay'。
        """
        df_shunfeng = outfile_df[outfile_df['trans_type'].isin(['shunfeng'])].copy()
        df_alipay = outfile_df[outfile_df['trans_type'] == 'alipay'].copy()
        df_ccb = outfile_df[outfile_df['trans_type'].isin(['CCB'])].copy()
        # 删除 'trans_type' 列
        df_shunfeng.drop(columns=['trans_type'], inplace=True)
        df_alipay.drop(columns=['trans_type'], inplace=True)
        df_ccb.drop(columns=['trans_type'], inplace=True)
        
        # 转换 'value_date' 列的日期格式
        for df in [df_shunfeng, df_alipay, df_ccb]:
            if 'value_date' in df.columns:
                df['value_date'] = pd.to_datetime(df['value_date'], errors='coerce').fillna(pd.Timestamp('today'))
                df['value_date'] = df['value_date'].dt.strftime('%Y%m%d')
                # df['value_date'] = pd.to_datetime(df['value_date']).dt.strftime('%Y%m%d')
     
        return df_shunfeng, df_alipay, df_ccb

    def out_to_template_file(self,region,outfile_df,file_type,mapping_list):
        """
        输出到文件
        """
        # 取得输出文件模板
        post_file_writer, out_file_path, item_name = self.file_util.open_template(
            region, file_type
        )
        # 打印数据到模板
        self.print_data_to_template(writer=post_file_writer, df=outfile_df,mapping_list=mapping_list)
        # 云存储
        self.move_file_to_cos(item_name, out_file_path)

    def out_to_template(self,region,outfile_df,file_type):
        """
        输出到格式化模板
        """
        # 加载Excel模板
        wb = openpyxl.load_workbook('Template.xltx')
        ws = wb.active

        # 将DataFrame转换为行
        rows = dataframe_to_rows(outfile_df, index=False, header=False)

        # 将行写入工作表
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 保存工作表为xlsx文件
        wb.template = False
        wb.save('File.xlsx')

    def get_post_data(self, region):
        """
        从bankstatement数据表中取得
        post文件数据
        """
        # and value_date = %s
        sql_query = """SELECT *,'' as ctype FROM bl_bank_statement 
                        WHERE otc_region = %s 
                        and ipf_status in ('user_confirmed','single_matched') 
                        order by trans_type,
                        CASE
                            WHEN otc_region = 'CN'
                                AND trans_type = 'alipay'
                                AND cn_sort_key ~ '^[0-9]+$'
                            THEN CAST(cn_sort_key AS INTEGER)
                            WHEN otc_region = 'CN'
                                AND trans_type != 'alipay' 
                            THEN NULL  -- 字符串情况下不转数字
                        END,
                        CASE
                            WHEN otc_region = 'CN' THEN cn_sort_key
                            ELSE NULL
                        END; 
                        """
        # and ipf_status = 'post_confirmed'
        # and bank_name = %s and currency = %s
        # and value_date = %s
        # and ipf_status = 'donepost'
        # parameters = (region,bank,currency,post_date,)
        parameters = (region,)
        rst = self.db.execute_query_col_name(sql_query, parameters)
        # print(rst)
        return rst

    def get_aloc_data(self, region):
        """
        销账可输出文件数据
        """
        sql_query = f"SELECT * FROM bl_aging_allocate WHERE otc_region = %s and aloc_status in ('user_confirmed','SO match') order by aloc_group "
        
        parameters = (region,)

        # # ========
        # sql_query = f"SELECT * FROM bl_aging_allocate WHERE otc_region = %s and aloc_status in (?, ?)"
        # parameters = (region, 'user_confirmed', 'so_matched')
        # # ========
        # aloc_statuses = ('user_confirmed', 'so_matched')
        # sql_query = f"SELECT * FROM bl_aging_allocate WHERE otc_region = %s and aloc_status in ({', '.join(['%s']*len(aloc_statuses))})"
        # parameters = (region,) + aloc_statuses

        rst = self.db.execute_query_col_name(sql_query, parameters)
        return rst

    def get_out_file_def(self, region, file_type):
        """
        取得配置数据
        """
        sql_query = "SELECT * FROM sc_output_file_define WHERE otc_region = %s and file_type = %s;"
        parameters = (region,file_type, )

        rst = self.db.execute_query_to_pandas(sql_query, parameters)

        return rst

    # def open_template(self, region):
    #     path = self.file_util.run_path('template')

    #     template_file = os.path.join(path,'Cash Application Tracker.xlsx')

    #     out_path = os.path.join(self.file_util.run_path('post_file'),region)

    #     if not os.path.exists(out_path):
    #         os.makedirs(out_path)

    #     out_file = os.path.join(out_path, 'Cash Application Tracker.xlsx')

    #     shutil.copyfile(template_file, out_file )

    #     writer = pd.ExcelWriter(out_file)

    #     return writer

    def print_data(self, region, bank, writer, df: pd.DataFrame,header=False):
        """
        输出数据到文件
        """
        try:
            df.to_excel(writer, sheet_name=region, startrow=0, header=header, index=False)
        except Exception as e:
            self.logger.error(f"Saving excel error : {e}")
        writer.save()

    def print_data_to_template(self, writer, df: pd.DataFrame, mapping_list):
        """
        输出数据到模板文件
        """
        headr_row = 3
        try:
            worksheet = writer.sheets['temp']
            for idx, row in df.iterrows():
                erow = idx + headr_row + 1
                for mapping in mapping_list:
                    col_dst = mapping[self.C_DST]
                    col_idx = mapping[self.C_IDX]
                    worksheet.cell(row=erow, column=int(col_idx), value=row[col_dst])
        except Exception as e:
            self.logger.error(f"Saving excel error : {e}")
        writer.save()


    def trans_bank_data_to_history(self,df):
        """
        银行数据导入银行历史数据表
        """
        table_name="bl_his_bank_statement"
        df["update_time"] = datetime.now()
        ret = self.db.upsert_table(df, table_name)
        return ret
    
    def update_bank_data(self,table_name,df):
        """
        已入账的银行数据，更新状态，复制 /转存 至历史表
        """
        # 更新入账完成数据状态
        self.update_post_status(table_name=table_name,df=df)
        # 完成数据转存到银行历史表
        self.trans_bank_data_to_history(df=df)
        # TODO 从银行数据删除已完成数据

    def update_post_status(self, table_name, df):
        """
        更新银行数据
        将数据标志为 已经输出文件
        """
        df["ipf_status"] = self.STS_PD
        
        # print(df)
        ret = self.db.upsert_table(df, table_name)
        return ret

    def update_aloc_data(self,table_name,df):
        """
        已入账的银行数据，更新状态，复制 /转存 至历史表
        """
        # 更新入账完成数据状态
        self.update_aloc_status(table_name=table_name,df=df)
        # 完成数据转存到银行历史表
        self.trans_aloc_data_to_history(df=df)

    def update_aloc_status(self, table_name, df):
        """
        更新银行数据
        将数据标志为 已经输出文件
        """
        df["aloc_status"] = self.STS_PD
        df["is_allocating"] = None
        ret = self.db.upsert_table(df, table_name)
        return ret
    
    def trans_aloc_data_to_history(self,df):
        """
        银行数据导入银行历史数据表
        """
        table_name="bl_his_allocate"
        ret = self.db.upsert_table(df, table_name)
        return ret
    
    def move_file_to_cos(self, item_name, out_file_path):
        """
        保存到云存储
        """
        if self.file_util.is_local == 1:
            return 
        cos, bucket = self.file_util.create_cos()
        self.file_util.multi_part_upload("ifp-post", item_name, out_file_path, cos)
        os.remove(out_file_path)

    def get_document_number(self, dest_df, region):
        """
        处理document number
        """
        rules = self.get_doc_num_rule(region)
        for ind_s, rule in rules.iterrows():
            work_type, target_dest_df, rest_dest_df = self.df_filter(dest_df, rule,region)
            if not target_dest_df.empty:
                latest_increase_no=""
                if work_type == "num":
                    target_dest_df, latest_increase_no = self.doc_num_increase(
                        target_dest_df, rule
                    )
                elif work_type == "date":
                    target_dest_df, latest_increase_no = self.doc_num_date(
                        target_dest_df, rule
                    )
                elif work_type == "SG":
                    filter_sg1 = dest_df["payment"].isin(["ADV"])
                    filter_sg2 = dest_df["sales_order"].str.len() == 8
                    target_dest_df = dest_df[filter_sg1 & filter_sg2].copy()
                    rest_dest_df = dest_df[~filter_sg1 | ~filter_sg2].copy()
                    latest_increase_no = rule["doc_number"]
                    if not target_dest_df.empty:
                        target_dest_df["document_no"] = target_dest_df.apply(
                            lambda x: self.handle_column("sales_order", x), axis=1
                        )
                    if not rest_dest_df.empty:
                        # work_type 是SG 数据的 payment 是 SCQ 不用increseno。
                        # 得到SCQ的 df scq_df，scq_df的document_no 设置为 从 asw_text 列提取出来的数字内容
                        # payment 是 SCQ 以外的部分，还是按照现在的做法
                        scq_df = rest_dest_df[rest_dest_df["doctype"] == "SCQ"].copy()
                        non_scq_df = rest_dest_df[rest_dest_df["doctype"] != "SCQ"].copy()
                        if not scq_df.empty:
                            scq_df["document_no"] = scq_df["asw_text"].apply(self.extract_digits_from_asw_text)

                        rest_dest_df, latest_increase_no = self.doc_num_increase(
                            rest_dest_df, rule
                        )
                        # 合并两个df到rest_dest_df
                        rest_dest_df = pd.concat([scq_df, non_scq_df])
                elif work_type == "CN":
                    filter_cn1 = dest_df["trans_type"].isin(
                        rule["cn_trans_type"].split(",")
                    )
                    target_dest_df = dest_df[filter_cn1].copy()
                    rest_dest_df = dest_df[~filter_cn1].copy()
                    if rule["cn_trans_type"] != "shunfeng":
                        closure_day_df = self.get_closure_day_df(
                            region, datetime.now().date()
                        )
                        target_dest_df, latest_increase_no = self.handle_cn_not_cod(
                            target_dest_df, rule, closure_day_df
                        )
                    else:
                        target_dest_df, latest_increase_no = self.handle_cn_cod(
                            target_dest_df, rule
                        )
                elif work_type =="amount" or work_type =="bank_ref" or work_type =="other":
                    if work_type =="amount" or work_type =="bank_ref":
                        if work_type=="amount":
                            target_dest_df['document_no'] = target_dest_df[work_type].apply(self.process_value)
                        else:
                            target_dest_df['document_no'] = target_dest_df.apply(self.assign_document_no, axis=1)
                    else:
                        
                        target_dest_df, latest_increase_no = self.doc_num_increase(
                            target_dest_df, rule
                        )
                else:
                    target_dest_df["document_no"] = target_dest_df.apply(
                        lambda x: self.handle_column(work_type, x), axis=1
                    )
                    latest_increase_no = ""
                rules.loc[ind_s, "doc_number"] = latest_increase_no
                dest_df = pd.concat([target_dest_df, rest_dest_df])
            else:
                continue
        return dest_df, rules

    def extract_digits_from_asw_text(self, text: str) -> str:
        """
        从 asw_text 中提取第一个连续的数字串作为 document_no
        """
        import re
        if pd.isna(text):
            return ""
        match = re.search(r"\d+", str(text))
        return match.group(0) if match else ""


    def assign_document_no(self, row):
        """
        判断是否数字
        """
        # 
        if self.extract_tran_num(row['narrative2'])=='175':
            return row['reference1']
        else:
            return self.process_value(row['amount'])
    
    def process_value(self,x):
        """
        转换为整数，社区小数部分
        """
        try:
            value = float(x)
            return int(round(value))
        except (ValueError, TypeError):
            return 0000
    
    def extract_tran_num(self,val):
        """
        提取括号中的数字
        """
        number=0
        if isinstance(val, (str, bytes)):
            match = re.search(r'\((\d+)\)', val)
            if match:
                number = match.group(1)
        return number
    
    listscq = ['175','399']
    lists5 = ['175','399','575']
    def set_doctype(self,df, listscq, lists5):
        """
        设定doctype   
        1. 175，399 是SCQ
        2. 175,399,575,以外，重复的数据 SBT
        3. 175,399,575,以外 非重复的 A60 
        """
        # 将ctype列转换为字符串类型，确保可以正确比较
        df['ctype'] = df['ctype'].astype(str).str.zfill(3)
        # 1. 175，399 是SCQ
        is_in_listscq = df['ctype'].isin(listscq)
        # 2. 175,399,575,以外，重复的数据 SBT
        # 3. 175,399,575,以外 非重复的 A60
        not_in_lists5 = ~df['ctype'].isin(lists5)
        # 计算cid的重复情况
        cid_duplicates = df['customer_id'].duplicated(keep=False)
        
        # 设置doctype列的初始值
        df['doctype'] = 'A60'
        # 根据条件设置doctype
        df.loc[is_in_listscq, 'doctype'] = 'SCQ'
        # 3. 175,399,575,以外 非重复的 A60
        df.loc[not_in_lists5 & cid_duplicates, 'doctype'] = 'SBT'
        df = df.sort_values(by='doctype')
        return df

    def get_customer_name_by_id(self, region, customer_id):
        """
        按 region 和 customer_id 从 bl_customer_master 取 customer_name。
        返回 None 表示未找到或查询出错。
        """
        try:
            sql_query = (
                "SELECT customer_name "
                "FROM bl_customer_master "
                "WHERE otc_region = %s AND customer_id = %s"
            )
            params = (region, customer_id)
            df = self.db.execute_query_to_pandas(sql_query, params)
            if df is None or df.empty:
                return None
            # 取第一条的 customer_name
            return df.loc[0, "customer_name"]
        except Exception as e:
            # 记录错误但不要抛出，让调用方继续处理其它行
            try:
                self.logger.error(f"get_customer_name_by_id error region={region} customer_id={customer_id}: {e}")
            except Exception:
                pass
            return None


    def fix_data_miss(self, df, region):
        """
        根据地区配置执行数据修复规则
        SG: 如果 asw_text 为伪空（NaN, 空串, "null", "nan", "none", « NULL » 等），
            则用 bl_customer_master 中对应 customer_id 的 customer_name 填充 asw_text。
        使用单条 SQL（region, customer_id）查询，带简易缓存避免重复查询。
        """
        # 必须列检查
        if 'asw_text' not in df.columns or 'customer_id' not in df.columns:
            return df

        import numpy as np

        def is_null_like(series):
            s = series.fillna("").astype(str).str.strip()
            simple_nulls = {"", "none", "nan", "null"}
            pattern = r'^[\"\'\u00ab\u00bb\u201c\u201d<>]*null[\"\'\u00ab\u00bb\u201c\u201d<>]*$'
            mask_simple = s.str.lower().isin(simple_nulls)
            mask_quoted = s.str.match(pattern, case=False, na=False)
            return series.isna() | mask_simple | mask_quoted

        if region == "SG":
            mask_to_fill = is_null_like(df['asw_text'])
            if not mask_to_fill.any():
                return df

            # 缓存 customer_id -> customer_name，避免重复 DB 查询
            customer_cache = {}
            filled_count = 0

            # 遍历需要填充的行
            for idx in df.loc[mask_to_fill].index:
                cust_id_raw = df.at[idx, 'customer_id']
                if cust_id_raw is None:
                    continue
                cust_id = str(cust_id_raw).strip()
                if cust_id == "":
                    continue

                # 先查缓存
                if cust_id in customer_cache:
                    name = customer_cache[cust_id]
                else:
                    # 单条查询
                    name = self.get_customer_name_by_id(region, cust_id)
                    customer_cache[cust_id] = name  # 可能是 None

                # 若找到了 name，则赋值
                if name:
                    df.at[idx, 'asw_text'] = name
                    filled_count += 1

            try:
                self.logger.info(f"dataMissFix: region={region}, filled asw_text for {filled_count} rows using bl_customer_master")
            except Exception:
                pass

        return df



    def set_df_doctype(self,df,region):
        """
        根据地区设置doctype
        IN地区：根据Transaction Type确定doctype
        其他地区：根据payment字段确定doctype
        """
        if region=="IN":
            df['ctype'] = df['narrative2'].apply(self.extract_tran_num)
            df = self.set_doctype(df, self.listscq, self.lists5)
        else:
            df = self.set_doctype_by_payment(df)
        
        return df

    def set_doctype_by_payment(self, df):
        """
        IN以外地区根据payment字段设置doctype
        如果payment值是ADV，doctype设置为SBT，其他设为SCQ
        """
        # 确保payment列存在
        if 'payment' not in df.columns:
            self.logger.warning("payment column not found, setting doctype to SCQ")
            df['doctype'] = 'SCQ'
            return df
        
        # 根据payment值设置doctype
        df['doctype'] = df['payment'].apply(lambda x: 'SBT' if x == 'ADV' else 'SCQ')
        
        return df


    def get_doc_num_rule(self, region):
        """
        获得document number规则
        """
        sql_query = f"SELECT * FROM sc_sap_document_define WHERE otc_region like %s "
        parameters = ("%" + region + "%",)
        rst = self.db.execute_query_col_name(sql_query, parameters)
        return rst

    def get_closure_day_df(self, region, current_date):
        """
        中国特殊处理获得开张日
        """
        sql_query = f"SELECT * FROM sc_month_open_date WHERE otc_region = %s and open_date = %s"
        parameters = (
            region,
            current_date,
        )
        rst = self.db.execute_query_col_name(sql_query, parameters)
        return rst

    def df_filter(self, dest_df, rule, region):
        """
        按照payment,currency过滤出该条规则的目标数据集
        """
        work_type = rule["work_type"]
        if region == "IN":
            filter1 = dest_df["doctype"].isin(rule["payment_type"].split(","))
        else:
            if rule["payment_type"]=="ALL":
                filter1=True
            else:
                filter1 = dest_df["payment"].isin(rule["payment_type"].split(","))
        filter2 = dest_df["currency"].isin(rule["currency"].split(","))
        target_dest_df = dest_df[filter1 & filter2].copy()
        # rest_dest_df = dest_df[~filter1 | ~filter2]
        rest_dest_df = dest_df[~(filter1&filter2)]
        return work_type, target_dest_df, rest_dest_df

    def doc_num_increase(self, target_dest_df, rule):
        """
        序列号自增处理
        """
        latest_increase_no = int(rule["doc_number"])
        for ind_s, row_s in target_dest_df.iterrows():
            latest_increase_no += 1
            target_dest_df.loc[ind_s, "document_no"] = latest_increase_no
        return target_dest_df, latest_increase_no

    def doc_num_date(self, target_dest_df, rule):
        """
        当前日期赋值处理
        """
        for ind_s, row_s in target_dest_df.iterrows():
            target_dest_df.loc[ind_s, "document_no"] = (
                str(datetime.now().date().strftime("%Y%m%d"))[2:]
            )
        return target_dest_df, datetime.now().date().strftime("%Y%m%d")

    def handle_cn_not_cod(self, target_dest_df, rule, closure_day_df):
        """
        中国非cod处理
        """
        if not closure_day_df.empty:
            increase_num = 0
            closure_month = closure_day_df.loc[0, "closure_month"]
        else:
            increase_num = int(rule["doc_number"][-4:])
            closure_month = rule["doc_number"][:6]
        return self.cn_df_iterrows(target_dest_df, increase_num, closure_month, 4)

    def handle_cn_cod(self, target_dest_df, rule):
        """
        中国cod处理
        """
        current_day = datetime.now().date().strftime("%Y%m%d")
        increase_num = (
            0
            if datetime.now().date().isoweekday() == 4
            else int(rule["doc_number"][-3:])
        )
        return self.cn_df_iterrows(target_dest_df, increase_num, current_day, 3)

    def cn_df_iterrows(self, target_dest_df, increase_num, concat_str, flag):
        """
        中国序列号自增处理
        """
        for ind_s, row_s in target_dest_df.iterrows():
            increase_num += 1
            target_dest_df.loc[ind_s, "document_no"] = str(
                concat_str + str(increase_num).zfill(flag)
            )
            if row_s["trans_type"] in ["CCB", "shunfeng"]:
                target_dest_df.loc[ind_s, "bank_code"] = "C01"
            elif row_s["trans_type"] == "alipay":
                target_dest_df.loc[ind_s, "bank_code"] = "A01"
        return target_dest_df, str(concat_str + str(increase_num).zfill(flag))

    def handle_column(self, column_name, x):
        """
        根据自身字段的共通处理
        """
        if column_name == "sales_order":
            if x[column_name]:
                if len(x[column_name]) == 8:
                    return x[column_name][-7:]
        elif column_name == "amount":
            return int(x[column_name])
        elif column_name == "bank_ref":
            return x[column_name] if x[column_name].isdigit() else int(x["amount"])
        
    def create_target_dataframe(self, src_df, mapping_list):
        '''
        DF to DF convert

        根据 raw DF to bl DF 的字段映射定义
        把 src_data rawDF 转换为 dest_df blDF

        '''

        dest_df = pd.DataFrame(dtype=object)

        add_cols = {}
        for mapping in mapping_list:
            col_src = mapping[self.C_SRC]
            col_dst = mapping[self.C_DST]

            # Check if source column exists in the source DataFrame
            if col_src in src_df.columns:
                # Copy the source column values to multiple destination columns dest_val1 = dest_val1.replace('\xa0', '')
                dest_df[col_dst] = src_df[col_src].copy().replace('\xa0', '')
            else:
                # 定义的源列名不在表中时，用列名赋值内容。
                if col_src.find("+") == -1:
                    dest_df[col_dst] = str(col_src)
                    add_cols[col_dst] = col_src
                else:
                    # dest_df[col_dst] = src_df.apply(lambda row: ''.join([str(row[col]) for col in col_src.split("+")]), axis=1)
                    dest_df[col_dst] = src_df.apply(lambda row: ''.join([str(row[col])[:20] if row[col] is not None else '' for col in col_src.split("+")]), axis=1)
                    add_cols[col_dst] = dest_df[col_dst]
        
        # 为什么要这样才能赋值成功？
        for key,value in add_cols.items():
            dest_df[key] = value.replace('\xa0', '')

        # self.logger.debug(dest_df)

        return dest_df
    
    def reassign_group_numbers(self,df, group_column):
        """
        重新映射组号
        从1开始
        """
        # 获取唯一的组号并排序
        unique_groups = sorted(df[group_column].unique())
        
        # 创建一个从1开始的新组号的映射
        group_mapping = {old_group: new_group for new_group, old_group in enumerate(unique_groups, start=1)}
        
        # 替换DataFrame中的组号
        df[group_column] = df[group_column].map(group_mapping)
        
        return df

    def raw_column_define(self, maping_str):
        '''
        解析列名到列名的定义映射
        到List Dict结构
        SG|otc_region;account_name|account_name;account_number|account_number;
        '''
        result_list = []
        for line in maping_str.split(';'):
            if line:
                parts = line.strip().split('|')
                col_index = 0
                col_src = ""
                col_dst = ""
                if len(parts) == 2:
                    col_src, col_dst = parts
                elif len(parts) == 3:
                    col_src, col_dst, col_index = parts
                
                col_info = {
                    self.C_SRC: col_src,
                    self.C_DST: col_dst,
                    self.C_IDX: col_index
                }
                result_list.append(col_info)
        return result_list
    
    def set_bank_code(self,df,region):
        """
        设定bank code

        """
        com_df = self.comm_util.get_comm_define(region)
        com_df.set_index("def_type", inplace=True)
        def_value_str = self.comm_util.get_com_def_by_name(com_df, "BANK_CODE")
        bank_code_map = json.loads(def_value_str)  # 转换为 dict

        if region=="CM":
            return df
        def get_bank_code(row):
            key = row['otc_region'] + row['bank_branch_name'] + row['currency']
            return bank_code_map.get(key, None)
        df['bank_code'] = df.apply(get_bank_code, axis=1)
        return df
    
    def get_summary_data(self, post_df):
        """按照 otc_region 和 currency 分组计算汇总数据"""
        summary_df = post_df.groupby(['otc_region', 'currency']).agg({
            'amount': ['sum', 'count']  # 对 amount 列同时计算合计和个数
        }).reset_index()
        
        # 重置列名
        summary_df.columns = ['otc_region', 'currency', 'amount', 'count']
        
        return summary_df

    def check_existing_record(self, otc_region, currency, work_date):
        """检查是否存在对应记录"""
        query = """
            SELECT post_amount, post_count 
            FROM bl_cash_sumdata 
            WHERE otc_region = %s 
            AND currency = %s 
            AND work_date = %s
        """
        parameters = (otc_region, currency, work_date)
        return self.db.execute_query_to_pandas(query, parameters)

    def update_existing_record(self, otc_region, currency, work_date, new_amount, new_count, existing_amount, existing_count):
        """更新已存在的记录"""
        update_query = """
            UPDATE bl_cash_sumdata 
            SET post_amount = %s,
                post_count = %s,
                updater = %s,
                update_time = %s
            WHERE otc_region = %s 
            AND currency = %s 
            AND work_date = %s
        """
        total_amount = new_amount + existing_amount
        total_count = new_count + int(existing_count)
        update_parameters = (
            total_amount,
            total_count,
            'postsystem',
            datetime.now(),
            otc_region,
            currency,
            work_date
        )
        self.db.execute_update_query(update_query, update_parameters)

    def insert_new_record(self, otc_region, currency, work_date, amount, count):
        """插入新记录"""
        insert_query = """
            INSERT INTO bl_cash_sumdata (
                otc_region, currency, work_date, 
                post_amount, post_count,
                creator, create_time, updater, update_time
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        current_time = datetime.now()
        insert_parameters = (
            otc_region,
            currency.strip(),
            work_date,
            amount,
            count,
            'postsystem',
            current_time,
            'postsystem',
            current_time
        )
        self.db.execute_insert_query(insert_query, insert_parameters)

    def update_cash_data(self, post_df):
        """更新现金数据的主方法"""
        # 计算汇总数据
        self.logger.debug(f"Summarize post amount and count start. ")
        summary_df = self.get_summary_data(post_df)
        current_date = datetime.now().date()
        
        # 处理每条汇总记录
        for _, row in summary_df.iterrows():
            otc_region = row['otc_region']
            currency = row['currency']
            amount_sum = row['amount']
            count_sum = row['count']
            
            # 检查记录是否存在
            existing_record = self.check_existing_record(otc_region, currency, current_date)
            
            if not existing_record.empty:
                # 更新已存在记录
                self.update_existing_record(
                    otc_region, 
                    currency, 
                    current_date, 
                    amount_sum,
                    count_sum,
                    existing_record['post_amount'].iloc[0],
                    existing_record['post_count'].iloc[0].astype(int)
                )
                self.logger.debug(f"Update post sum amount and count. {otc_region}, {currency}")
            else:
                # 插入新记录
                self.insert_new_record(otc_region, currency, current_date, amount_sum, count_sum)
                self.logger.debug(f"Insert post sum amount and count. {otc_region}, {currency}")
        self.logger.debug(f"Summarize post amount and count done. {otc_region}")
        
    def get_unallocated_data(self, otc_region):
        """获取未分配的数据"""
        unaloc_sql = """
        SELECT otc_region, currency, amount, doc_type 
        FROM bl_aging_allocate 
        WHERE otc_region = %s 
        AND (aloc_status NOT IN ('user_confirmed', 'SO match', 'data_posted') or aloc_status is null)
        AND doc_type NOT IN %s
        """
        unaloc_df = self.db.execute_query_to_pandas(unaloc_sql, (otc_region, self.doc_types))
        
        # 按otc_region和currency分组计算未分配金额和件数
        unaloc_grouped = unaloc_df.groupby(['otc_region', 'currency']).agg({
            'amount': 'sum',
            'doc_type': 'count'  # 使用otc_region列来计数记录数
        }).reset_index()
        
        return unaloc_grouped.rename(columns={
            'amount': 'unaloc_amt',
            'doc_type': 'unaloc_count'
        })

    def get_existing_records(self, otc_region, current_date):
        """获取现有的记录"""
        check_sql = """
        SELECT otc_region, currency, allocate_amount, allocate_count
        FROM bl_cash_sumdata
        WHERE otc_region = %s
        AND work_date = %s
        """
        return self.db.execute_query_to_pandas(check_sql, (otc_region, current_date))

    def update_existing_alo_record(self, row, existing_record, updater, current_date):
        """更新现有记录"""
        update_sql = """
        UPDATE bl_cash_sumdata
        SET allocate_amount = %s,
            allocate_count = %s,
            undone_amount = %s,
            undone_count = %s,
            updater = %s,
            update_time = %s
        WHERE otc_region = %s
        AND currency = %s
        AND work_date = %s
        """
        new_allocate_amount = abs(row['amount']) + abs(existing_record['allocate_amount'])
        new_allocate_count = row['count'] + existing_record['allocate_count']
        
        parameters = (
            abs(float(new_allocate_amount)),
            abs(int(new_allocate_count)),
            abs(float(row['unaloc_amt'])),
            abs(int(row['unaloc_count'])),
            updater,
            datetime.now(),
            row['otc_region'],
            row['currency'].strip(),
            current_date
        )
        self.db.execute_update_query(update_sql, parameters)

    def insert_new_alo_record(self, row, current_date):
        """插入新记录"""
        insert_sql = """
        INSERT INTO bl_cash_sumdata (
            otc_region, currency, work_date, 
            allocate_amount, allocate_count,
            undone_amount, undone_count,
            creator, create_time, updater, update_time
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        curr_time = datetime.now()
        parameters = (
            row['otc_region'],
            row['currency'].strip(),
            current_date,
            abs(row['amount']),
            abs(row['count']),
            abs(row['unaloc_amt']),
            abs(row['unaloc_count']),
            'system',
            curr_time,
            'system',
            curr_time
        )
        self.db.execute_insert_query(insert_sql, parameters)

    def process_and_update_cash_data(self, dest_df, otc_region):
        """主处理方法"""
        
        self.logger.debug(f"Summarize allocate amount start. {otc_region} ")
        aloc_df = dest_df.copy()
        
        # 1. 获取未分配数据
        unaloc_grouped = self.get_unallocated_data(otc_region)
        updater = "alocsystem"
        
        # 2. 获取当前日期的现有记录
        current_date = datetime.now().date()
        existing_records = self.get_existing_records(otc_region, current_date)
        
        # 3. 过滤并按otc_region和currency分组统计金额和件数
        filtered_df = aloc_df[~aloc_df['doc_type'].isin(list(self.doc_types))]
        aloc_df = filtered_df.groupby(['otc_region', 'currency']).agg({
            'amount': 'sum',
            'doc_type': 'count'  # 使用doc_type列来计数记录数
        }).reset_index()
        aloc_df = aloc_df.rename(columns={'doc_type': 'count'})

        # 4. 合并已分配和未分配数据
        merged_df = pd.merge(aloc_df, unaloc_grouped, on=['otc_region', 'currency'], how='outer')
        merged_df['amount'] = merged_df['amount'].fillna(0)
        merged_df['count'] = merged_df['count'].fillna(0)
        merged_df['unaloc_amt'] = merged_df['unaloc_amt'].fillna(0)
        merged_df['unaloc_count'] = merged_df['unaloc_count'].fillna(0)
        
        # 5. 处理每条记录
        for _, row in merged_df.iterrows():
            existing = existing_records[
                (existing_records['otc_region'] == row['otc_region']) & 
                (existing_records['currency'] == (row['currency'].strip()))
            ]
            
            if not existing.empty:
                # 更新现有记录
                self.update_existing_alo_record(row, existing.iloc[0], updater, current_date)
                self.logger.debug(f"update aloc sum amount. {otc_region} ,{row['currency'].strip()}")
            else:
                # 插入新记录
                self.insert_new_alo_record(row, current_date)
                self.logger.debug(f"Insert aloc sum amount. {otc_region},{row['currency'].strip()} ")

        self.logger.debug(f"Summarize allocate amount done. {otc_region} ")
        return True

    def create_alloc_header_data(self, dest_df, region, action_user):
        """
        生成 Allocation Header sheet 数据
        """
        # Allocation Date: 取 aloc_date 第一条，格式 DDMMYY
        alloc_date = datetime.now().strftime("%d%m%y")

        # has_aloc_column = "aloc_date" in dest_df.columns
        # has_nonnull_rows = dest_df["aloc_date"].notna().any()
        
        # if has_nonnull_rows:
        if (
            "aloc_date" in dest_df.columns
                and not dest_df.empty
                and dest_df["aloc_date"].notna().any()
            ):

            first_date = dest_df["aloc_date"].dropna().iloc[0]
            alloc_date = pd.to_datetime(first_date).strftime("%d%m%y")

        # Region
        region_val = dest_df["otc_region"].iloc[0] if "otc_region" in dest_df.columns else region
        # Group Count
        group_count = dest_df["aloc_group"].nunique() if "aloc_group" in dest_df.columns else 0
        # User
        user = action_user
        header_data = [{
            "Allocation Date": alloc_date,
            "Region": region_val,
            "Group Count": group_count,
            "User": user
        }]
        return pd.DataFrame(header_data)

    def create_alloc_detail_data(self, dest_df, mapping_list):
        """
        生成 Allocation Detail sheet 数据，严格按 mapping_list 定义
        """
        detail_df = pd.DataFrame()
        for mapping in mapping_list:
            src = mapping[self.C_SRC]
            dst = mapping[self.C_DST]
            if src and src in dest_df.columns:
                detail_df[dst] = dest_df[src]
            elif src == "0":
                detail_df[dst] = 0
            else:
                detail_df[dst] = ""
        # 特殊处理：Due Date 格式化
        if "Due Date" in detail_df.columns:
            detail_df["Due Date"] = pd.to_datetime(detail_df["Due Date"], errors="coerce").dt.strftime("%Y%m%d")
        # 排序：customer_id, aloc_group, amount, doc_type 升序
        sort_cols = []
        for col in ["Customer ID", "Group No", "Original Amount", "Document type"]:
            if col in detail_df.columns:
                sort_cols.append(col)
        if sort_cols:
            detail_df = detail_df.sort_values(by=sort_cols, ascending=True, kind='mergesort')
        
        # 调用新增的差额处理逻辑
        self.set_small_difference_by_group(detail_df)

        return detail_df

    def set_small_difference_by_group(self, detail_df):
        """
        按 Group No 分组，对 Original Amount 求和并处理 Small Difference。
        如果分组合计金额不为0，且本组所有 Small Difference 为0，
        则将绝对值金额填入该组所有记录的 Small Difference。
        特殊规则：
            当 Comments 中包含 "Partial"（不区分大小写）时，
            Small Difference 一律置为 0，不做处理。
        """
        required_cols = ["Group No", "Original Amount", "Small Difference", "Comments"]
        if not all(col in detail_df.columns for col in required_cols):
            return

        # 先统一处理包含 Partial 的行 -> Small Difference 设为 0
        mask_partial = detail_df["Comments"].str.contains("partial", case=False, na=False)
        detail_df.loc[mask_partial, "Small Difference"] = 0

        # 按 Group No 分组求 Original Amount 的绝对值和
        group_sums = detail_df.groupby("Group No")["Original Amount"].sum().round(2).abs()

        for group_no, abs_sum in group_sums.items():
            if abs_sum != 0:
                group_mask = detail_df["Group No"] == group_no

                # 如果该组里含有 Partial，就跳过，不修改（保证 Small Difference=0）
                if detail_df.loc[group_mask, "Comments"].str.contains("partial", case=False, na=False).any():
                    continue

                # 将 abs_sum 设置为该组所有记录的 Small Difference
                detail_df.loc[group_mask, "Small Difference"] = abs_sum



    def out_to_alloc_template_file(self, region, header_df, detail_df, file_type):
        """
        输出 Allocation 数据到模板文件，包含 Header 和 Detail 两个 sheet
        """
        writer, out_file_path, item_name = self.file_util.open_template(region, file_type, filename="Allocation")
        try:
            header_df.to_excel(writer, sheet_name="Allocation Header", startrow=0, header=True, index=False)
            detail_df.to_excel(writer, sheet_name="Allocation Detail", startrow=0, header=True, index=False)
            self.logger.info(f"Allocation template file created successfully. Header rows: {len(header_df)}, Detail rows: {len(detail_df)}")
        except Exception as e:
            self.logger.error(f"Error creating allocation template file: {e}")
            raise CustomException(f"Error creating allocation template file: {e}")
        finally:
            writer.save()
        self.move_file_to_cos(item_name, out_file_path)

    def get_common_cid(self, region):
        """
        取得规则表定义，如果有记录，返回 div_rule1 中的数字部分，否则返回空字符串
        例如 div_rule1 为 "const:123456:::rem"，返回 "123456"
        """
        rname = "set_constant_" + region.lower()
        if region == "MY":
            rname = "set_constant_" + region.lower() + "_hsbc"

        sql_query = """
            SELECT * 
            FROM sc_search_rule 
            WHERE rule_seq = 1 AND rule_name = %s  
            ORDER BY group_order, rule_seq;
        """
        parameters = (rname,)

        rst = self.db.execute_query_to_pandas(sql_query, parameters)

        if not rst.empty and "div_rule1" in rst.columns:
            div_rule1 = rst["div_rule1"].iloc[0] or ""
            # 提取以 "const:" 开头后面的数字部分
            match = re.search(r"const:(\d+)", div_rule1)
            if match:
                return match.group(1)
        return ""

    def output_post_data_cn_csv(self, region, action_user="system"):
        """
        CN地区post数据输出为csv，按trans_type分文件，文件名带账期和自增号
        """
        self.logger.info(f"CN Post CSV file output start Region: {region} ")
        try:
            bz_date = self.comm_util.get_daytime_string()
            post_action_name = f"{region}_post_file_output"
            self.comm_util.update_file_outputing_status(
                region=region,
                action_name=post_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["6"],
                bl_message=f"Post file output action on Region:{region} start.",
                ac_user=action_user,
            )

            table_name = "bl_bank_statement"

            # 取得post对象数据
            dest_df = self.get_post_data(region)
            if dest_df.empty:
                self.logger.error("There is no data to post")
                return

            # 处理document number
            dest_df, define_table_df = self.get_document_number(dest_df, region)
            # 设定bankcode
            dest_df = self.set_bank_code(dest_df, region)

            # 取得输出字段顺序
            def_df = self.get_out_file_def(region, self.POST_FILE)
            out_fields = def_df.loc[0, "out_field_define"]
            fields_list = [f for f in out_fields.split("|") if f]

            # 只保留需要的列
            # dest_df = dest_df[fields_list]

            # 按trans_type分文件
            df_shunfeng, df_alipay, df_ccb = self.split_dataframe(dest_df)
            # 对拆分后的 DataFrame 再次裁剪列，防止分组逻辑中新增其他列
            df_shunfeng = df_shunfeng[[col for col in fields_list if col in df_shunfeng.columns]]
            df_alipay   = df_alipay[[col for col in fields_list if col in df_alipay.columns]]
            df_ccb      = df_ccb[[col for col in fields_list if col in df_ccb.columns]]

            file_dfs = [("CCB", df_ccb), ("alipay", df_alipay), ("shunfeng", df_shunfeng)]

            # 账期月和自增号逻辑
            today = datetime.now().date()
            closure_df = self.get_closure_day_df(region, today)
            if not closure_df.empty:
                # 今天是开账日
                account_month = closure_df["account_month"].iloc[0]
                seq = 1
            else:
                # 不是开账日，取昨天的文件名
                # yesterday = today - pd.Timedelta(days=1)
                com_df = self.comm_util.get_comm_define(region)
                com_df.set_index("def_type", inplace=True)
                last_file = self.comm_util.get_com_def_by_name(com_df, "cn_posting")
                if last_file and len(last_file) >= 8:
                    account_month = last_file[:6]
                    seq = int(last_file[6:8]) + 1
                else:
                    # fallback
                    account_month = today.strftime("%Y%m")
                    seq = 1

            # 依次输出3个文件
            for idx, (ftype, df) in enumerate(file_dfs):
                if df.empty:
                    continue
                file_seq = seq + idx
                file_name = f"{account_month}{file_seq:02d}.csv"
                out_path = self.file_util.run_path("post_file")
                out_path = os.path.join(out_path,region)
                if not os.path.exists(out_path):
                    os.makedirs(out_path)
                out_file = os.path.join(out_path, file_name)
                df.to_csv(out_file, index=False, header=False, encoding="utf-8-sig")
                self.logger.info(f"CN Post CSV file output: {out_file}")
                # 上传到cos
                self.move_file_to_cos(file_name, out_file)
                # 只用最后一个文件名更新sc_com_define
                if idx == 2 or (idx < 2 and all(d.empty for _, d in file_dfs[idx+1:])):
                    file_base_name = os.path.splitext(file_name)[0]  # 去掉 .csv 后缀
                    self.update_cn_posting_file_name(region, file_base_name)

            # 后续数据处理
            dest_df = dest_df.fillna(
                {"credit_amount": 0, "debit_amount": 0, "update_time": datetime.now()}
            )
            # 更新数据标识
            # self.update_post_status(table_name, dest_df)
            # 更新状态，转存到历史表 
            self.update_bank_data(table_name, dest_df)

            # 最新的document number更新到DB
            self.db.upsert_table(define_table_df, "sc_sap_document_define")
            
            # 保存入账结果数据，报表用
            self.update_cash_data(dest_df)

            self.comm_util.update_file_outputed_status(
                region=region,
                action_name=post_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["7"],
                bl_message=f"Post file output action on Region:{region} finished.",
                ac_user=action_user,
            )
        except Exception as e:
            self.logger.error(str(e))
            self.logger.error(traceback.format_exc())
            self.comm_util.update_file_outputed_status(
                region=region,
                action_name=post_action_name,
                business_date=bz_date,
                status=self.comm_util.BL_STS["11"],
                bl_message=f"Post file output  action on Region:{region} Failed .",
                ac_user=action_user,
            )
        self.logger.info(f"CN Post CSV file output Finished. Region: {region} ")

    def update_cn_posting_file_name(self, region, file_name):
        """
        更新sc_com_define表的def_context字段为最新文件名
        """
        sql = """
            UPDATE sc_com_define
            SET def_context = %s
            WHERE otc_region = %s AND def_type = 'cn_posting'
        """
        params = (file_name, region)
        self.db.execute_update_query(sql, params)